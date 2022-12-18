import codecs
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os
import sys
from collections import defaultdict

dir_parse = {}
COLOR_MAP = [
    'FFFFFF',
    'F5B090',
    'FCD7A1',
    'FFF9B1',
    'D7E7AF',
    'A5D4AD',
    'A2D7D4',
    '9FD9F6',
    'A3BCE2',
    'A59ACA',
    'CFA7CD',
    'F4B4D0',
    'F5B2B2'
]

def parse_dirs(target_dir, target_file):
    for dr in os.listdir(target_dir):
        p = os.path.join(target_dir, dr, target_file)
        result_dic = parse_dir(p)
        dir_parse[dr] = result_dic

def parse_dir(pfile) -> dict:
    if not os.path.exists(pfile):
        return
    result_dic = defaultdict(list)
    with codecs.open(pfile, 'r', 'utf-8') as f:
        for line in f:
            parse_tpl = parse_line(line.rstrip())
            # print(parse_tpl)
            if parse_tpl:
                result_dic[parse_tpl[0]].append(parse_tpl[1])
    print(result_dic)
    return result_dic

def parse_line(line: str):
    if not line.strip():
        return None
    if line.strip()[:1] == '#':
        return None
    key, val = line.split('=', maxsplit=1)
    return (key, val)

def print_csv(outputfile: str):
    all_keys = sorted(list({key for dic in dir_parse.values() for key in dic.keys()}))
    colums = sorted(dir_parse.keys())
    with open(outputfile, 'w') as f:
        w = csv.writer(f, delimiter='\t')
        w.writerow(['key'] + colums)
        for key in all_keys:
            row = [key]
            for cl in colums:
                val = dir_parse[cl][key]
                if (len(val) == 0):
                    row.append('')
                elif (len(val) > 1):
                    # リストのまま追加
                    row.append(val)
                else:
                    row.append(val[0])
            w.writerow(row)

def print_excel(outputfile: str):
    all_keys = sorted(list({key for dic in dir_parse.values() for key in dic.keys()}))
    colums = sorted(dir_parse.keys())
    wb = Workbook()
    ws = wb.active
    rownum = 1
    print_worksheet_row(ws, rownum, ['key'] + colums)
    for key in all_keys:
        rownum += 1
        row = [key]
        for cl in colums:
            val = dir_parse[cl][key]
            if (len(val) == 0):
                row.append('')
            elif (len(val) > 1):
                # リストのまま追加
                row.append(str(val))
            else:
                row.append(val[0])
        print_worksheet_row(ws, rownum, row)
    wb.save(outputfile)

def print_worksheet_row(ws, rownum: int, vals: list):
    color_dic = defaultdict(str)
    color_idx = 0
    for idx, val in enumerate(vals):
        if not color_dic[val]:
            color_dic[val] = COLOR_MAP[color_idx]
            color_idx += 1
        ws.cell(rownum, idx+1, val)
        ws.cell(rownum, idx+1).fill = PatternFill(patternType='solid', fgColor=color_dic[val])
    print(color_dic)

if __name__ == '__main__':
    arg = sys.argv[1:]
    if (len(arg) != 3):
        print('引数の数が不正です')
        sys.exit()
    target_dir = arg[0]
    target_file = arg[1]
    output_file = arg[2]
    parse_dirs(target_dir, target_file)
    # print_csv(output_file)
    print_excel(output_file)
